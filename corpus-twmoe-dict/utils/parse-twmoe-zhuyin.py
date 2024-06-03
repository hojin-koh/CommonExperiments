#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2020-2024, Hojin Koh
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Parse twmoe dictionary, supposedly from https://language.moe.gov.tw/001/Upload/Files/site_content/M0001/respub/dict_concised_download.html

import sys
import re
import unicodedata

import openpyxl

def main():
    objSheet = openpyxl.load_workbook(sys.argv[1]).worksheets[0]
    sWords = set()
    for i in range(2, objSheet.max_row+1):
        plist = objSheet.cell(i, 7).value
        if not plist or len(plist) < 2: continue
        plist.replace('　', ' ')
        plist = re.sub(R'(ˊ)(.)|(ˇ)(.)|(ˋ)(.)', R'\1 \2', plist)
        for w in plist.strip().split():
            w = w.strip().removeprefix('˙')
            w = w.strip().removesuffix('ㄦ')
            if len(w) < 2: continue
            sWords.add(w)

            w = w.removesuffix('ˊ').removesuffix('ˇ').removesuffix('ˋ')
            if len(w) < 2: continue
            sWords.add(w)

    for w in sorted(sWords):
        print(w)

if __name__ == '__main__':
    main()
