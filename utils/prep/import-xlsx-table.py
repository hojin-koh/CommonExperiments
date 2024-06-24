#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2020-2024, Hojin Koh
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Import table data from xlsx format source data

import sys

import openpyxl
from opencc import OpenCC

def main():
    fileExcel = sys.argv[1]
    idWorksheet = int(sys.argv[2])-1
    rowBegin = int(sys.argv[3])
    objConv = None
    if len(sys.argv) > 4 and sys.argv[4].endswith(".json"):
        objConv = OpenCC(sys.argv[4])
        del sys.argv[4]
    cols = tuple((int(v) for v in sys.argv[4:]))

    objSheet = openpyxl.load_workbook(fileExcel).worksheets[idWorksheet]
    for row in range(rowBegin, objSheet.max_row+1):
        aThis = []
        for col in cols:
            w = objSheet.cell(row, col).value.strip()
            if not w or len(w) < 1: continue
            if objConv:
                w = objConv.convert(w)
            aThis.append(w)

        if len(aThis) != len(cols): continue
        print("\t".join(aThis))

if __name__ == '__main__':
    main()
