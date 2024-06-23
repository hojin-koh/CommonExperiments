#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2020-2024, Hojin Koh
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Parse twmoe dictionary, supposedly from https://language.moe.gov.tw/001/Upload/Files/site_content/M0001/respub/dict_concised_download.html

import re
import sys
import unicodedata

import openpyxl

def main():
    objSheet = openpyxl.load_workbook(sys.argv[1]).worksheets[0]
    for i in range(2, objSheet.max_row+1):
        w = objSheet.cell(i, 1).value.strip()
        if not w or len(w) < 2: continue
        # Delete punctuations
        w = "".join(c for c in w if not unicodedata.category(c).startswith('P'))

        # Blacklist
        if w.endswith("縣") and w != "知縣": continue
        if len(w) > 6 and re.search(R"多一事不如少一事|建設|戰機|打狗|系統|地址|號誌", w): continue
        # Long proper nouns
        if len(w) > 6 and re.search(R"指數$|獎$|制度$|政策$|中心$", w): continue

        # Number words or special words
        if re.match(R"^[一二三四五六七八九]{3}[^一二三四五六七八九十]", w):
            print(w[:3])
        if w.find("天安門") != -1:
            print("天安門")
            print("六四")
        if w.find("奧林匹克") != -1:
            print("奧林匹克")
        if w.find("奧斯卡") != -1:
            print("奧斯卡")

        # Post-processing some weird words
        elif len(w) == 8:
            print(w[:4])
            print(w[4:])
        elif len(w) == 10:
            print(w[:5])
            print(w[5:])

        print(w)

if __name__ == '__main__':
    main()
